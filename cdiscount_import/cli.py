import sys
import pathlib
import configparser
import pandas as pd
import numpy as np
import openpyxl
import re
import os
import requests
from openpyxl.utils.dataframe import dataframe_to_rows
import pprint
import json
import plenty_api


PROG_NAME = 'cdiscount_import'
USER = str(os.getlogin())
if sys.platform == 'linux':
    BASE_PATH = pathlib.Path('/') / 'home' / USER / '.config'
elif sys.platform == 'win32':
    BASE_PATH = pathlib.Path('C:\\') / 'Users' / USER / '.config'

if not BASE_PATH.exists():
    pathlib.Path(BASE_PATH).mkdir(parents=True, exist_ok=True)

CONFIG_FOLDER = BASE_PATH / PROG_NAME
if not CONFIG_FOLDER.exists():
    pathlib.Path(CONFIG_FOLDER).mkdir(parents=True, exist_ok=True)

CONFIG_PATH = CONFIG_FOLDER / 'config.ini'

if not CONFIG_PATH.exists():
    open(CONFIG_PATH, 'a').close()


# Constants

MAX_LONG_DESC_LEN = 5000
MAX_LONG_LABEL_LEN = 132
MAX_SHORT_DESC_LEN = 420
MAX_SHORT_LABEL_LEN = 30
MAX_SELLER_REF_LEN = 50
MAX_PARENT_SKU_LEN = 50
MAX_EAN_LEN = 13
MAX_MARKET_COLOR_LEN = 50
ID_INDEX = 0


cdiscount_list = [
    "Seller product ref", "Barcode", "Brand", "Product nature",
    "Category code", "Short product label", "Long product label",
    "Product description", "Image 1" , "Family sku", "Size (limited)",
    "Marketing colour", "Product's marketing description", "Image 2",
    "Image 3", "Image 4", "MFPN", "Sous-état", "Licence", "Licences",
    "Type de Produit", "Gamme", "Description du produit", "Collection",
    "Modèles", "Composition du lot", "Lavable", "Pays d'origine",
    "Certifications et normes", "Plus produit", "Forme", "Confort",
    "Mentions légales", "Produit adapté à", "Type de fabrication",
    "Type de rideau", "Extensible", "Composition", "Type de fibre",
    "Type de tissu", "Garnissage", "Enveloppe", "Déhoussable",
    "Densité de tissage", "Grammage", "Finitions", "Fermeture",
    "Type de chaleur", "Couleur principale", "Couleur(s)", "Motif",
    "Motifs", "Style", "Utilisation", "Dimensions", "Dimensions du linge",
    "Dimensions bonnet", "Usage unique", "Type d'attaches", "Type de pièce",
    "Traitement de protection", "Conseil d'entretien",
    "Conseils d'utilisation",
    "Dimensions brutes - article emballé (L x l x H)", "Poids emballé",
    "Poids net", "Garantie (²)", "Garantie additionnelle", "Observations",
    "durée de disponibilité des pièces détachées essentielles à l’utilisation du produit",
    "Notes", "Labels et certifications"
]


class InvalidConfig(Exception):
    """
    Exception raised when the provided configuration is not valid.

    Attributes:
            section     -   The section that contains an invalid option
            option      -   The option which is invalid
            message     -   Explanation of the error
    """
    def __init__(self, section: str, option: str = ''):
        self.section = section
        self.option = option
        super().__init__()

    def __str__(self):
        """Build the exception message from the different attributes."""
        if not self.option:
            return f"missing section [{self.section}]"
        return f"missing option `{self.option}` in section [{self.section}]"


class PlentyFetch:
    def __init__(self, config: configparser.ConfigParser,
                 debug: bool = False) -> None:
        self.config = config
        self.__check_config()
        self.debug = debug
        self.referrer_id = int(self.config['plenty']['referrer_id'])
        self.attribute_mapping = {}
        self.manufacturers = []
        self.variations = []
        self.item_ids = {}
        self.errors = []

    def __check_config(self):
        """
        Check if the configuration contains all required sections and options.
        """
        required_options = {
            'plenty': [
                'base_url', 'color_attribute_id', 'size_attribute_id',
                'referrer_id', 'ean_barcode_id', 'plenty_id'
            ], 'category_mapping': []}

        for section in required_options:
            if not self.config.has_section(section=section):
                raise InvalidConfig(section=section)

            for option in required_options[section]:
                if not self.config.has_option(section=section, option=option):
                    raise InvalidConfig(section=section, option=option)

    def connect(self):
        """Connect to the plentyAPI"""
        self.api = plenty_api.PlentyApi(
            base_url=self.config['plenty']['base_url'],
            use_keyring=True,
            debug=self.debug
        )

    def __get_market_mapping(self, attribute_id: int, market_id: int) -> dict:
        """
        Get a attribute mapping for a specifc marketplace from Plentymarkets.

        Parameters:
            attribute_id[int]   -   ID assigned by Plentymarkets for the
                                    attribute
            market_id   [int]   -   ID assinged by Plentymarkets for the
                                    marketplace
        """
        maps = requests.get(
            self.api.url + '/rest/items/attributes/values/maps',
            headers=self.api.creds
        ).json()
        entries = maps['entries']

        page = 1
        while (page < maps['lastPageNumber']):
            page += 1
            new_maps = requests.get(
                self.api.url +
                f'/rest/items/attributes/values/maps?page={page}',
                headers=self.api.creds
            ).json()
            entries += new_maps['entries']

        return {
            str(entry['attributeValueId']): entry['marketInformation1']
            for entry in entries
            if entry['attributeId'] == attribute_id
            and entry['marketId'] == market_id
        }

    def __get_attribute_mappings(self, lang: str) -> dict:
        """
        Create a map of attribute names for size and color attribute values.

        Parameters:
            lang        [str]   -   2 letter abbr. of the target language

        Return:
                        [dict]
        """
        attributes = self.api.plenty_api_get_attributes(additional=['values'])
        color_id = int(self.config['plenty']['color_attribute_id'])
        size_id = int(self.config['plenty']['size_attribute_id'])
        cdiscount_mappings = self.__get_market_mapping(
            attribute_id=color_id, market_id=self.referrer_id)
        if not cdiscount_mappings:
            raise RuntimeError("No mapped color values for Cdiscount")

        mapping = {'color': cdiscount_mappings}
        for attribute in attributes:
            if attribute['id'] == size_id:
                mapping['size'] = {
                    str(value['id']):str(name['name'])
                    for value in attribute['values']
                    for name in value['valueNames'] if name['lang'] == lang
                }

        return mapping

    def __get_color_attribute(self, variation: dict) -> str:
        """
        Get the markting color from the color mapping table for a color.

        Parameters:
            variation   [dict]  -   JSON of a single variation from the
                                    Plentymarkets REST API

        Return:
                        [str]   -   Name of the size
        """
        try:
            attributes = variation['variationAttributeValues']
        except KeyError:
            return 'No color attribute found'

        plenty_id = int(self.config['plenty']['color_attribute_id'])
        for attribute in attributes:
            if attribute['attributeId'] == plenty_id:
                color_id = str(attribute['attributeValue']['id'])
                try:
                    return self.attribute_mapping['color'][color_id]
                except KeyError:
                    return 'No color mapping found'

        return 'No color attribute found'

    def __get_size_attribute(self, variation: dict) -> str:
        """
        Get the size name from the size mapping table for the given attribute.

        Parameters:
            variation   [dict]  -   JSON of a single variation from the
                                    Plentymarkets REST API

        Return:
                        [str]   -   Name of the size
        """
        try:
            attributes = variation['variationAttributeValues']
        except KeyError:
            return ''

        plenty_id = int(self.config['plenty']['size_attribute_id'])
        for attribute in attributes:
            if attribute['attributeId'] == plenty_id:
                value_id = str(attribute['attributeValue']['id'])
                try:
                    return self.attribute_mapping['size'][value_id]
                except KeyError:
                    return ''

        return ''

    def __get_size_property(self, variation: dict, lang: str) -> str:
        """
        As alternative to the size attribute, get the size from a property.

        Some products don't have a size attribute as they are one-size products
        Cdiscount requires a size name, so get the size from a property as
        alternative.

        Parameters:
            variation   [dict]  -   JSON of a single variation from the
                                    Plentymarkets REST API
            lang        [str]   -   2 letter abbr. of the target language

        Return:
                        [str]
        """
        if not self.config.has_option(section='plenty',
                                      option='size_property_id'):
            return ''

        try:
            properties = variation['variationProperties']
        except KeyError:
            return ''

        prop_id = int(self.config['plenty']['size_property_id'])
        for prop in properties:
            if prop['propertyId'] == prop_id:
                for name in prop['names']:
                    if lang.lower() == name['lang'].lower():
                        return name['value']

        return ''

    def __get_barcode(self, variation: dict) -> str:
        """
        Get the 13 character EAN (GTIN-13) barcode from Plentymarkets.

        Parameters:
            variation   [dict]  -   JSON of a single variation from the
                                    Plentymarkets REST API

        Return:
                        [str]
        """
        try:
            barcodes = variation['variationBarcodes']
        except KeyError:
            return 'No barcode found'

        barcode_id = int(self.config['plenty']['ean_barcode_id'])
        for barcode in barcodes:
            if barcode['barcodeId'] == barcode_id:
                if len(barcode['code']) != MAX_EAN_LEN:
                    return 'Invalid EAN barcode length'
                return barcode['code']

        return 'No EAN barcode found'

    def __get_category(self, variation: dict) -> str:
        """
        Get the default category ID for the mandant that runs Cdiscount and
        map it to the valid Cdiscount category ID.

        Parameters:
            variation   [dict]  -   JSON of a single variation from the
                                    Plentymarkets REST API

        Return:
                        [str]
        """
        try:
            categories = variation['variationDefaultCategory']
        except KeyError:
            return 'No category found'

        plenty_id = int(self.config['plenty']['plenty_id'])
        for category in categories:
            if category['plentyId'] == plenty_id:
                category_id = str(category['branchId'])
            else:
                continue
            try:
                return self.config['category_mapping'][category_id]
            except KeyError:
                return 'No mapped cdiscount category'
        return 'No category for mandant'

    def __get_brand(self, variation: dict) -> str:
        """
        Get brand name from the manufacturer list of Plentymarkets.

        Parameters:
            variation   [dict]  -   JSON of a single variation from the
                                    Plentymarkets REST API

        Return:
                        [str]
        """
        try:
            item = variation['item']
        except KeyError:
            return 'No item found'

        if not self.manufacturers:
            self.manufacturers = self.api.plenty_api_get_manufacturers()
        for manufacturer in self.manufacturers:
            # As we fetched the ID from the item, it is guranteed that we find
            # a match
            if manufacturer['id'] == item['manufacturerId']:
                return manufacturer['name']

    def __get_images(self, variation : dict) -> list:
        """
        Get a maximum of 4 images for the Cdiscount columns.

        The last image should always be a swatch image (an image that
        represents multiple variations at once).

        Parameters:
            variation   [dict]  -   JSON of a single variation from the
                                    Plentymarkets REST API

        Return:
                        [str]
        """
        try:
            images = variation['images']
        except KeyError:
            return []

        image_list = []
        for image in images:
            for availability in image['availabilities']:
                if availability['value'] == self.referrer_id:
                    image_list.append(
                        {'url': image['url'], 'position': image['position']}
                    )

        if not image_list:
            return []

        image_list = sorted(image_list, key=lambda item: item.get('position'))
        swatch_images = []
        for index, image in enumerate(image_list):
            # This is a highly specific condition for our use case
            if re.search('swatch', image['url'].lower()):
                swatch_images.append(image_list.pop(index))

        if len(image_list) > 4 and len(swatch_images) == 0:
            image_list = image_list[:4]
        elif len(image_list) > 3 and len(swatch_images) >= 1:
            image_list = image_list[:3] + [swatch_images[0]]
        else:
            image_list += swatch_images[:4-len(image_list)]

        return [x['url'] for x in image_list]

    def extract_data(self):
        """
        Get all the variations from the API that have the referrerId of
        cdiscount.  Then cycle through the json file for the data that is
        needed and do checks if they fulfill cdiscounts requirements and put
        them into a list of lists.
        """
        self.attribute_mapping = self.__get_attribute_mappings(lang='fr')
        variations = self.api.plenty_api_get_variations(
            refine = {'referrerId': self.referrer_id}, additional = [
                'variationProperties', 'variationBarcodes',
                'variationDefaultCategory', 'images',
                'variationAttributeValues', 'parent', 'item'
            ],
            lang='fr'
        )

        image_block = []
        img = False
        for variation in variations:
            err = False

            try:
                self.item_ids[str(variation['itemId'])].append(variation['id'])
            except KeyError:
                self.item_ids[str(variation['itemId'])] = [variation['id']]

            if variation['isMain'] == True:
                continue

            marketing_color = self.__get_color_attribute(variation=variation)
            if marketing_color in  ['No color attribute found',
                                    'No color mapping found']:
                err = True

            size = self.__get_size_attribute(variation=variation)
            if not size:
                size = self.__get_size_property(variation=variation, lang='fr')
            if size == '':
                err = True
                size = 'Empty Value'

            barcode = self.__get_barcode(variation=variation)
            if barcode in ['No barcode found', 'No EAN barcode found',
                           'Invalid EAN barcode length']:
                err = True

            try:
                parent_sku = variation['parent']['number']
                if len(parent_sku) > MAX_PARENT_SKU_LEN:
                    parent_sku = 'Too long'
                    err = True
            except IndexError:
                parent_sku = 'Not Found'
                err = True

            if parent_sku == '':
                err = True
                parent_sku = 'Empty Value'

            try:
                seller_ref = str(variation['id'])
                if len(seller_ref) > MAX_SELLER_REF_LEN:
                    seller_ref = 'Too long'
                    err = True
            except IndexError:
                seller_ref = 'Not Found'
                err = True

            if seller_ref == '':
                err = True
                seller_ref = 'Empty Value'

            brand = self.__get_brand(variation=variation)
            if brand == 'No item found':
                err = True

            category_id = self.__get_category(variation=variation)
            if category_id in ['No category found', 'No category for mandant',
                               'No mapped cdiscount category']:
                err = True

            product_nature = 'Standard'

            image_block = self.__get_images(variation=variation)
            if not image_block:
                err = True
                image_block = ['No Image found']

            data = [
                seller_ref, barcode, brand, product_nature, category_id,
                image_block[0], parent_sku, size, marketing_color
            ] + image_block[1:]
            if err:
                self.errors.append(data)
                err = False
                continue

            self.variations.append(data)

    def get_texts(self):
        """
        Get all the parents from the variations which have been extracted in
        extract_data(). Then cycle through the json file for the text data that
        is needed and do checks if they fulfill cdiscounts requirements and put
        them into a list of lists  and after put them into the item list
        created by extract_data().
        """
        item_string_list = "','".join(self.item_ids.keys())

        items = (self.api.plenty_api_get_items(
            refine={'id':item_string_list}, lang='fr'
        ))

        texts = []
        error_texts = []
        for item in items:
            err = False
            item_id = str(item['id'])

            # We only pull the french texts therefore we are guranteed to get
            # the right text at index 0.
            if item['texts']:
                text = item['texts'][0]
                if len(text['description']) <= MAX_LONG_DESC_LEN:
                    long_description = text['description']
                else:
                    err = True
                    long_description = 'Text too long'

                if len(text['name1']) <= MAX_SHORT_LABEL_LEN:
                    short_label = text['name1']
                else:
                    err = True
                    short_label = 'Text too long'

                if len(text['name2']) <= MAX_LONG_LABEL_LEN:
                    long_label = text['name2']
                else:
                    err = True
                    long_label = 'Text too long'

                if len(text['shortDescription']) <= MAX_SHORT_DESC_LEN:
                    short_description = text['shortDescription']
                else:
                    err = True
                    short_description = 'Text too long'

                data = {
                    'item_id': item_id,
                    'short_label': short_label,
                    'long_label': long_label,
                    'short_description': short_description,
                    'long_description': long_description
                }
            else:
                data = {
                    'item_id': item_id,
                    'short_label': 'No french text found',
                    'long_label': 'No french text found',
                    'short_description': 'No french text found',
                    'long_description': 'No french text found'
                }
                err = True

            if err:
                error_texts.append(data)
                err = False
            else:
                texts.append(data)

        count_list = []
        for error in error_texts:
            for count, variation in enumerate(self.variations):
                if int(variation[ID_INDEX]) in self.item_ids[error['item_id']]:
                    self.errors.append(variation + [x for x in error.values()])
                    count_list.append(count)
        count_list.reverse()

        for i in count_list:
            self.variations.pop(i)

        for text in texts:
            for variation in self.variations:
                if int(variation[ID_INDEX]) in self.item_ids[text['item_id']]:
                    variation.insert(5, text['short_label'])
                    variation.insert(6, text['long_label'])
                    variation.insert(7, text['short_description'])
                    variation.insert(12, text['long_description'])


class CdiscountWriter:
    def __init__(self, filename: str, error_filename: str,
                 base_path: str = ''):
        if not base_path:
            base_path = pathlib.Path('.')
        else:
            base_path = pathlib.Path(base_path)

        self.filename = base_path / filename
        self.error_filename = base_path / error_filename

    def write_xlsx(self, variations: list):
        """
        Put the extracted data into a dataframe and write into an excel file.

        Parameters:
            variations  [list]  -   Extracted variations from the plentymarkets
                                    API in the correct order for the Cdiscount
                                    file
        """
        df = pd.DataFrame(variations)
        if len(df.index) == 0:
            pprint.pprint("No extracted variations found.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Linge de maison - rideau - store'
        ws['A1'] = 'Model:'
        ws['B1'] = 'Linge de maison - rideau - store'
        ws.append(['',''])
        ws.append(cdiscount_list)
        ws.append(['',''])
        for row in dataframe_to_rows(df, index = False, header = False):
            ws.append(row)

        wb.save(filename=self.filename)

    def write_error(self, errors: list):
        """
        Put the detected errors into a dataframe and write into an excel file.

        Parameters:
            errors      [list]  -   Detected errors while reading variations
                                    from the REST API
        """
        df = pd.DataFrame(errors)
        if len(df.index) == 0:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Linge de maison - rideau - store'
        ws['A1'] = 'Model:'
        ws['B1'] = 'Linge de maison - rideau - store'
        ws.append(['',''])
        ws.append([
            'seller_ref', 'barcode', 'brand', 'product_nature',
            'category_id', 'image_1', 'parent_sku', 'size', 'marketing_color',
            'image_2', 'short_label', 'long_label', 'short_description',
            'long_description'
        ])
        ws.append(['',''])
        for row in dataframe_to_rows(df, index = False, header = False):
            ws.append(row)

        wb.save(filename=self.error_filename)


def main():
    config = configparser.ConfigParser()
    config.read(CONFIG_PATH)

    base_path = ''
    if config.has_section(section='general'):
        if config.has_option(section='general', option='file_destination'):
            base_path = config['general']['file_destination']

    try:
        plenty_fetch = PlentyFetch(config=config)
    except InvalidConfig as err:
        pprint.pprint(f"{err}")
        sys.exit(1)

    cdiscount_writer = CdiscountWriter(filename='cdiscount_import.xlsm',
                                       error_filename='cdiscount_errors.xlsm',
                                       base_path=base_path)

    plenty_fetch.connect()
    plenty_fetch.extract_data()
    plenty_fetch.get_texts()
    cdiscount_writer.write_xlsx(variations=plenty_fetch.variations)
    cdiscount_writer.write_error(errors=plenty_fetch.errors)
